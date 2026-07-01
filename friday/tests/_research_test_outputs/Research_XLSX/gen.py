import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------------
# 1. Generate sample product data (₹ under 2000)
# ------------------------------------------------------------------
products = [
    # Brand, Model, Price(INR), Battery(hrs), Rating(0/1=percent), Weight(g), BT Version
    ("boAt", "Rockerz 450",     1499, 15, 4.2/5, 150, "5.0"),
    ("Realme", "Buds Wireless 2", 1999, 12, 4.5/5, 120, "5.0"),
    ("JBL", "Tune 110BT",       1799, 8,  4.0/5, 140, "5.0"),
    ("Noise", "Shots X4",       1299, 20, 4.1/5, 130, "5.1"),
    ("Boult Audio", "ProBass",  1199, 10, 3.9/5, 160, "5.0"),
    ("Mivi", "DuoPods M100",    1599, 24, 4.3/5, 110, "5.2"),
    ("pTron", "Bassbuds",       799,  8,  3.8/5, 100, "5.0"),
    ("Zebronics", "Zeal Elite", 1299, 12, 4.0/5, 145, "5.0"),
    ("Sony", "WI-C100",         1990, 10, 4.1/5, 150, "5.0"),
    ("OnePlus", "Nord Buds",    1899, 12, 4.4/5,  90, "5.2"),
]

# ------------------------------------------------------------------
# 2. Create workbook and sheets
# ------------------------------------------------------------------
wb = openpyxl.Workbook()
ws_dash = wb.active
ws_dash.title = "Dashboard"
ws_products = wb.create_sheet("Products")
ws_charts = wb.create_sheet("Charts")

# ------------------------------------------------------------------
# 3. Style definitions
# ------------------------------------------------------------------
header_fill = PatternFill(start_color="378ADD", end_color="378ADD", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
alt_fill_1 = PatternFill(start_color="E6F1FB", end_color="E6F1FB", fill_type="solid")
alt_fill_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
kpi_font = Font(bold=True, size=14)
label_font = Font(bold=True, size=10)
totals_font = Font(bold=True, size=11)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ------------------------------------------------------------------
# 4. PRODUCTS SHEET
# ------------------------------------------------------------------
headers = ["Brand", "Model", "Price", "Battery Life", "Rating", "Weight", "Bluetooth Version"]
row = 1
for col_idx, h in enumerate(headers, 1):
    cell = ws_products.cell(row=row, column=col_idx, value=h)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# fill data
for i, (brand, model, price, batt, rating, weight, bt) in enumerate(products):
    row = i + 2
    ws_products.cell(row=row, column=1, value=brand).font = Font(size=11)
    ws_products.cell(row=row, column=2, value=model).font = Font(size=11)
    
    price_cell = ws_products.cell(row=row, column=3, value=price)
    price_cell.number_format = '₹#,##0'
    price_cell.font = Font(size=11)
    
    batt_cell = ws_products.cell(row=row, column=4, value=batt)
    batt_cell.number_format = '0.0"h"'
    batt_cell.font = Font(size=11)
    
    rating_cell = ws_products.cell(row=row, column=5, value=rating)
    rating_cell.number_format = '0%'
    rating_cell.font = Font(size=11)
    
    weight_cell = ws_products.cell(row=row, column=6, value=weight)
    weight_cell.number_format = '0"g"'
    weight_cell.font = Font(size=11)
    
    ws_products.cell(row=row, column=7, value=bt).font = Font(size=11)
    
    # alternating row fill
    fill = alt_fill_1 if i % 2 == 0 else alt_fill_2
    for col in range(1, 8):
        ws_products.cell(row=row, column=col).fill = fill
        ws_products.cell(row=row, column=col).alignment = Alignment(horizontal='center', vertical='center')
        ws_products.cell(row=row, column=col).border = thin_border

# auto-filter on header
ws_products.auto_filter.ref = f"A1:G{len(products)+1}"

# freeze top row
ws_products.freeze_panes = "A2"

# auto-width columns for Products sheet
for col in range(1, 8):
    max_len = 0
    for row in range(1, len(products)+2):
        cell = ws_products.cell(row=row, column=col)
        if cell.value:
            length = len(str(cell.value))
            max_len = max(max_len, length)
    ws_products.column_dimensions[get_column_letter(col)].width = max_len + 3

# ------------------------------------------------------------------
# 5. DASHBOARD SHEET
# ------------------------------------------------------------------
# 5.1 KPI Row (row1)
kpi_data = [
    ("Total Products", f"=COUNTA(Products!A2:A{len(products)+1})"),
    ("Average Price", f"=AVERAGE(Products!C2:C{len(products)+1})"),
    ("Top Rating", f"=MAX(Products!E2:E{len(products)+1})"),
    ("Cheapest Option", f"=MIN(Products!C2:C{len(products)+1})"),
]

# place KPIs in row1, with labels in column A, C, E, G and values in B, D, F, H
col_offset = 0
for idx, (label, formula) in enumerate(kpi_data):
    col_label = 1 + col_offset
    col_value = 2 + col_offset
    # label
    cell_lbl = ws_dash.cell(row=1, column=col_label, value=label)
    cell_lbl.font = label_font
    cell_lbl.alignment = Alignment(horizontal='right', vertical='center')
    # value
    cell_val = ws_dash.cell(row=1, column=col_value, value=formula)
    cell_val.font = kpi_font
    cell_val.alignment = Alignment(horizontal='center', vertical='center')
    # format appropriately
    if "Price" in label:
        cell_val.number_format = '₹#,##0'
    elif "Rating" in label:
        cell_val.number_format = '0%'
    elif "Products" in label:
        cell_val.number_format = '0'
    col_offset += 2

# 5.2 Summary Table (product list with totals row)
start_row = 4  # row where table starts
table_headers = ["Brand", "Model", "Price", "Rating"]
for col_idx, h in enumerate(table_headers, 1):
    cell = ws_dash.cell(row=start_row, column=col_idx, value=h)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# Data rows using formulas pointing to Products sheet
num_products = len(products)
for i in range(num_products):
    row = start_row + 1 + i
    # Brand (col A)
    ws_dash.cell(row=row, column=1, value=f'=IF(Products!A{i+2}<>"",