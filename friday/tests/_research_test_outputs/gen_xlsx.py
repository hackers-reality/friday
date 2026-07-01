import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation

# =============================================================================
# Sample product data derived from research links and typical specifications
# =============================================================================
products = [
    {
        "Product": "realme Buds Air 3",
        "Brand": "realme",
        "Price": 1999,
        "Battery Life": "20 hrs",
        "Rating": 4.3,
        "Bluetooth Version": "5.3",
        "Warranty": "1 Year",
        "Buy Link": "https://buy.realme.com/in/goods/729"
    },
    {
        "Product": "boAt Rockerz 480",
        "Brand": "boAt",
        "Price": 1499,
        "Battery Life": "20 hrs",
        "Rating": 4.1,
        "Bluetooth Version": "5.0",
        "Warranty": "1 Year",
        "Buy Link": "https://www.boat-lifestyle.com/products/rockerz-480-wireless-headphone"
    },
    {
        "Product": "Noise Buds VS104",
        "Brand": "Noise",
        "Price": 1299,
        "Battery Life": "24 hrs",
        "Rating": 4.2,
        "Bluetooth Version": "5.2",
        "Warranty": "1 Year",
        "Buy Link": "https://www.gonoise.com/collections/earbuds-under-2000"
    },
    {
        "Product": "Boult Audio ProBass",
        "Brand": "Boult",
        "Price": 999,
        "Battery Life": "18 hrs",
        "Rating": 4.0,
        "Bluetooth Version": "5.0",
        "Warranty": "6 Months",
        "Buy Link": "https://goboult.co.in/collections/over-ear-bluetooth-headphone-with-mic"
    },
    {
        "Product": "Portronics Harmonics",
        "Brand": "Portronics",
        "Price": 799,
        "Battery Life": "20 hrs",
        "Rating": 3.8,
        "Bluetooth Version": "5.0",
        "Warranty": "1 Year",
        "Buy Link": "https://www.flipkart.com/headphones/portronics~brand/pr?sid=0pm,fcn,gc3"
    },
    {
        "Product": "realme Buds Q2",
        "Brand": "realme",
        "Price": 1599,
        "Battery Life": "16 hrs",
        "Rating": 4.2,
        "Bluetooth Version": "5.1",
        "Warranty": "1 Year",
        "Buy Link": "https://buy.realme.com/in/goods/760"
    },
    {
        "Product": "boAt Rockerz 550",
        "Brand": "boAt",
        "Price": 1799,
        "Battery Life": "30 hrs",
        "Rating": 4.3,
        "Bluetooth Version": "5.3",
        "Warranty": "1 Year",
        "Buy Link": "https://www.boat-lifestyle.com/products/boat-rockerz-plus-550"
    },
    {
        "Product": "Snapods View LCD TWS",
        "Brand": "Snapods",
        "Price": 1499,
        "Battery Life": "24 hrs",
        "Rating": 4.0,
        "Bluetooth Version": "5.2",
        "Warranty": "1 Year",
        "Buy Link": "https://www.snapup.life/collections/snapods-bluetooth-earbuds/products/snapods-view-lcd-display-tws-earbuds"
    },
    {
        "Product": "Snapods E320 Neo Gaming",
        "Brand": "Snapods",
        "Price": 999,
        "Battery Life": "20 hrs",
        "Rating": 3.9,
        "Bluetooth Version": "5.1",
        "Warranty": "1 Year",
        "Buy Link": "https://www.snapup.life/collections/snapods-bluetooth-earbuds/products/snapods-e320-neo-gaming-earbud"
    }
]

# =============================================================================
# Create workbook and sheets
# =============================================================================
wb = Workbook()
ws1 = wb.active
ws1.title = "Product Comparison"
ws2 = wb.create_sheet("Price Analysis")

# =============================================================================
# Helper styles
# =============================================================================
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

even_row_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")  # light blue
odd_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# =============================================================================
# Write "Product Comparison" sheet
# =============================================================================
headers1 = ["Product", "Brand", "Price", "Battery Life", "Rating", "Bluetooth Version", "Warranty", "Buy Link"]

# Write header row
for col_idx, header in enumerate(headers1, 1):
    cell = ws1.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Write product data
for row_idx, prod in enumerate(products, 2):
    row_data = [
        prod["Product"],
        prod["Brand"],
        prod["Price"],
        prod["Battery Life"],
        prod["