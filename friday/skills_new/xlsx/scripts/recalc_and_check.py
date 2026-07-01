"""
friday/skills/xlsx/scripts/recalc_and_check.py

Recalculates all formulas in a workbook (preferring real Excel via xlwings
if available, falling back to LibreOffice headless), then scans every cell
for Excel error values and prints a pass/fail summary.

Usage:
    python recalc_and_check.py output.xlsx
"""
import subprocess
import sys
from pathlib import Path

ERROR_VALUES = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!"}


def recalc_with_excel(path: Path) -> bool:
    try:
        import xlwings as xw
    except ImportError:
        return False
    try:
        app = xw.App(visible=False)
        app.display_alerts = False
        app.screen_updating = False
        try:
            wb = app.books.open(str(path.resolve()))
            wb.save()
            wb.close()
        finally:
            app.quit()
        print("Recalculated via MS Excel (xlwings).")
        return True
    except Exception as e:
        print(f"Excel COM recalculation failed ({e}), will try LibreOffice fallback.")
        return False


def recalc_with_libreoffice(path: Path) -> bool:
    import shutil
    if not shutil.which("soffice"):
        print("soffice not found on PATH — cannot recalculate. Run check_env.py.")
        return False
    out_dir = path.parent / "_recalc_tmp"
    out_dir.mkdir(exist_ok=True)
    subprocess.run(
        ["soffice", "--headless", "--convert-to", "xlsx", "--outdir", str(out_dir), str(path)],
        check=True,
    )
    recalced = out_dir / path.name
    if recalced.exists():
        recalced.replace(path)
        print("Recalculated via LibreOffice headless.")
        return True
    print("LibreOffice conversion did not produce expected output.")
    return False


def scan_for_errors(path: Path):
    from openpyxl import load_workbook
    wb = load_workbook(str(path), data_only=True)
    errors = []
    total_cells_checked = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                total_cells_checked += 1
                if isinstance(cell.value, str) and cell.value in ERROR_VALUES:
                    errors.append(f"{ws.title}!{cell.coordinate}: {cell.value}")
    print(f"Scanned {total_cells_checked} cells across {len(wb.worksheets)} sheet(s).")
    return errors


def main():
    if len(sys.argv) != 2:
        print(__doc__)
        sys.exit(1)

    path = Path(sys.argv[1])
    if not path.exists():
        print(f"File not found: {path}")
        sys.exit(1)

    recalced = recalc_with_excel(path) or recalc_with_libreoffice(path)
    if not recalced:
        print("WARNING: could not recalculate — error scan below reflects stale/unrecalculated values.")

    errors = scan_for_errors(path)

    print()
    if errors:
        print(f"FORMULA ERRORS FOUND ({len(errors)}):")
        for e in errors:
            print(f"  - {e}")
        sys.exit(1)
    else:
        print("No formula errors found.")
        sys.exit(0)


if __name__ == "__main__":
    main()
